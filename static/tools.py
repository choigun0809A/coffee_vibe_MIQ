import firebase_admin 
import os, json
from firebase_admin import credentials, firestore
import random as rand
from openpyxl import Workbook
from openpyxl.styles import Border, Side

from io import BytesIO

key_dict = json.loads(os.environ['KEY'])
cred = credentials.Certificate(key_dict)
firebase_admin.initialize_app(cred)
# type of intelligence: {question: answer}

db = firestore.client()

def generate_random_id():
    return rand.randint(1, 9999999)
    

def upload_c(val, val2):
    id = generate_random_id()
    db.collection('Choistests').document(str(id)).set(val)
    db.collection('Choistests_intelligences').document(str(id)).set(val2)
def upload_d(val, val2):
    id = generate_random_id()
    db.collection('Ochirstests').document(str(id)).set(val)
    db.collection('Ochirstests_intelligences').document(str(id)).set(val2)

def get_users_c():
    """Retrieve all documents from the 'intell' collection as a list."""
    users_ref = db.collection('Choistests')
    docs = users_ref.stream()
    chars = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")

    yy = 1

    wb = Workbook()
    ws = wb.active

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws["A" + str(yy)] = "Choistests answers"
    
    ws["A" + str(yy)].border = thin_border


    ws["H" + str(yy)] = f"sum of kids:"
    ws["H" + str(yy)].border = thin_border
    ws["I" + str(yy)] = "intelligence:"
    ws["I" + str(yy)].border = thin_border
    ws["J" + str(yy)] = f"question:"
    ws["J" + str(yy)].border = thin_border
    ws["K" + str(yy)] = "-2:"
    ws["K" + str(yy)].border = thin_border
    ws["L" + str(yy)] = f"-1:"
    ws["L" + str(yy)].border = thin_border
    ws["M" + str(yy)] = "0:"
    ws["M" + str(yy)].border = thin_border
    ws["N" + str(yy)] = f"1:"
    ws["N" + str(yy)].border = thin_border
    ws["O" + str(yy)] = "2:"
    ws["O" + str(yy)].border = thin_border

    all_kids = 0
    
    
    

    stats: dict[str, dict[int, int]] = {}
    
    for doc in docs:
        all_kids += 1
        name = doc.id
        data: dict[str, dict[str, int]] = doc.to_dict()
        ws["B" + str(yy)] = f"id:"
        ws["B" + str(yy)].border = thin_border
        ws["C" + str(yy)] = "intelligence:"
        ws["C" + str(yy)].border = thin_border
        ws["D" + str(yy)] = f"question:"
        ws["D" + str(yy)].border = thin_border
        ws["E" + str(yy)] = "answers:"
        ws["E" + str(yy)].border = thin_border

        yy += 1


        ws["B" + str(yy)] = name
        ws["B" + str(yy)].border = thin_border

        # type of intelligence: {question: answer}
        
        for type_of in data.keys():
            
            for question, answer in data[type_of].items():
                
                ws["B" + str(yy)].border = thin_border
                ws["C" + str(yy)] = type_of
                ws["C" + str(yy)].border = thin_border
                ws["D" + str(yy)] = f"{question}"
                ws["D" + str(yy)].border = thin_border
                ws["E" + str(yy)] = f"{answer}"
                ws["E" + str(yy)].border = thin_border
                try:
                    stats[question][answer] += 1
                except:
                    stats[question] = {-2: 0, -1: 0, 0: 0, 1: 0, 2: 0}
                    stats[question][answer] += 1
                yy += 1
    
    
    print(stats)

    yy = 2
    ws["H" + str(yy)] = all_kids
    ws["H" + str(yy)].border = thin_border
    docs = users_ref.stream()
    for doc in docs:
        data: dict[str, dict[str, int]] = doc.to_dict()
        for type_of in data.keys():
            ws["I" + str(yy)] = type_of
            ws["I" + str(yy)].border = thin_border

            for question, answer in data[type_of].items():
                print("yes")
                ws["J" + str(yy)] = question
                ws["J" + str(yy)].border = thin_border
                ws["K" + str(yy)] = str(round(stats[question][-2] / all_kids * 100, 2)) + "%"
                ws["K" + str(yy)].border = thin_border
                ws["L" + str(yy)] = str(round(stats[question][-1] / all_kids * 100, 2)) + "%"
                ws["L" + str(yy)].border = thin_border
                ws["M" + str(yy)] = str(round(stats[question][0] / all_kids * 100, 2)) + "%"
                ws["M" + str(yy)].border = thin_border
                ws["N" + str(yy)] = str(round(stats[question][1] / all_kids * 100, 2)) + "%"
                ws["N" + str(yy)].border = thin_border
                ws["O" + str(yy)] = str(round(stats[question][2] / all_kids * 100, 2)) + "%"
                ws["O" + str(yy)].border = thin_border

                ws["I" + str(yy)].border = thin_border
                yy += 1
            yy += 1
    users_ref = db.collection('Ochirstests_intelligences')
    ddocs = users_ref.stream()
    yy = 1
    for ddoc in ddocs:
        data: dict[str, float] = ddoc.to_dict()
        ws["P" + str(yy)] = f"id:"
        ws["P" + str(yy)].border = thin_border
        ws["Q" + str(yy)] = "Байгалийн:"
        ws["Q" + str(yy)].border = thin_border
        ws["R" + str(yy)] = f"Хөгжмийн:"
        ws["R" + str(yy)].border = thin_border
        ws["S" + str(yy)] = "Логик:"
        ws["S" + str(yy)].border = thin_border
        ws["T" + str(yy)] = f"Экстенциал:"
        ws["T" + str(yy)].border = thin_border
        ws["U" + str(yy)] = "Харилцааны:"
        ws["U" + str(yy)].border = thin_border
        ws["V" + str(yy)] = f"Хөдөлгөөний:"
        ws["V" + str(yy)].border = thin_border
        ws["W" + str(yy)] = "Үгийн:"
        ws["W" + str(yy)].border = thin_border
        ws["X" + str(yy)] = "Өөртөө чиглэсэн:"
        ws["X" + str(yy)].border = thin_border
        ws["Y" + str(yy)] = "Дүрслэх:"
        ws["Y" + str(yy)].border = thin_border
        yy+=1

        ws["P" + str(yy)] = f"{ddoc.id}"
        ws["P" + str(yy)].border = thin_border

        index = chars.index("Q")
        for _, percentage in data.items():
            ws[chars[index] + str(yy)] = f"{percentage}%"
            ws[chars[index] + str(yy)].border = thin_border
            index += 1
        yy += 2



    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream
        




        
    
    



    



    return None


def get_users_d():
    """Retrieve all documents from the 'intell' collection as a list."""
    users_ref = db.collection('Ochirstests')
    docs = users_ref.stream()
    
    chars = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")

    yy = 1

    wb = Workbook()
    ws = wb.active

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws["A" + str(yy)] = "Ochirstests answers"
    
    ws["A" + str(yy)].border = thin_border


    ws["H" + str(yy)] = f"sum of kids:"
    ws["H" + str(yy)].border = thin_border
    ws["I" + str(yy)] = "intelligence:"
    ws["I" + str(yy)].border = thin_border
    ws["J" + str(yy)] = f"question:"
    ws["J" + str(yy)].border = thin_border
    ws["K" + str(yy)] = "1:"
    ws["K" + str(yy)].border = thin_border
    ws["L" + str(yy)] = "0:"
    ws["L" + str(yy)].border = thin_border

    all_kids = 0
    
    
    

    stats: dict[str, dict[int, int]] = {}
    
    for doc in docs:
        all_kids += 1
        name = doc.id
        data: dict[str, dict[str, int]] = doc.to_dict()
        ws["B" + str(yy)] = f"id:"
        ws["B" + str(yy)].border = thin_border
        ws["C" + str(yy)] = "intelligence:"
        ws["C" + str(yy)].border = thin_border
        ws["D" + str(yy)] = f"question:"
        ws["D" + str(yy)].border = thin_border
        ws["E" + str(yy)] = "answers:"
        ws["E" + str(yy)].border = thin_border

        yy += 1


        ws["B" + str(yy)] = name
        ws["B" + str(yy)].border = thin_border


        # type of intelligence: {question: answer}
        
        for type_of in data.keys():
            
            for question, answer in data[type_of].items():
                
                ws["B" + str(yy)].border = thin_border
                ws["C" + str(yy)] = type_of
                ws["C" + str(yy)].border = thin_border
                ws["D" + str(yy)] = f"{question}"
                ws["D" + str(yy)].border = thin_border
                ws["E" + str(yy)] = f"{answer}"
                ws["E" + str(yy)].border = thin_border

                try:
                    stats[question][answer] += 1
                except:
                    stats[question] = {1: 0, 0: 0}
                    stats[question][answer] += 1
                yy += 1
    
    
    print(stats)

    yy = 2
    ws["H" + str(yy)] = all_kids
    ws["H" + str(yy)].border = thin_border
    docs = users_ref.stream()
    for doc in docs:
        data: dict[str, dict[str, int]] = doc.to_dict()
        for type_of in data.keys():
            ws["I" + str(yy)] = type_of
            ws["I" + str(yy)].border = thin_border

            # ws["H" + str(yy)] = f"sum of kids:"
            # ws["H" + str(yy)].border = thin_border
            # ws["I" + str(yy)] = "intelligence:"
            # ws["I" + str(yy)].border = thin_border
            # ws["J" + str(yy)] = f"question:"
            # ws["J" + str(yy)].border = thin_border
            # ws["K" + str(yy)] = "1:"
            # ws["K" + str(yy)].border = thin_border
            # ws["L" + str(yy)] = "0:"
            # ws["L" + str(yy)].border = thin_border
            for question, answer in data[type_of].items():
                print("yes")
                ws["J" + str(yy)] = question
                ws["J" + str(yy)].border = thin_border
                ws["K" + str(yy)] = str(round(stats[question][1] / all_kids * 100, 2)) + "%"
                ws["K" + str(yy)].border = thin_border
                ws["L" + str(yy)] = str(round(stats[question][0] / all_kids * 100, 2)) + "%"
                ws["L" + str(yy)].border = thin_border
                

                ws["I" + str(yy)].border = thin_border
                yy += 1
            yy += 1

    

    users_ref = db.collection('Ochirstests_intelligences')
    ddocs = users_ref.stream()
    yy = 1
    for ddoc in ddocs:
        data: dict[str, float] = ddoc.to_dict()
        ws["P" + str(yy)] = f"id:"
        ws["P" + str(yy)].border = thin_border
        ws["Q" + str(yy)] = "Байгалийн:"
        ws["Q" + str(yy)].border = thin_border
        ws["R" + str(yy)] = f"Хөгжмийн:"
        ws["R" + str(yy)].border = thin_border
        ws["S" + str(yy)] = "Логик:"
        ws["S" + str(yy)].border = thin_border
        ws["T" + str(yy)] = f"Экстенциал:"
        ws["T" + str(yy)].border = thin_border
        ws["U" + str(yy)] = "Харилцааны:"
        ws["U" + str(yy)].border = thin_border
        ws["V" + str(yy)] = f"Хөдөлгөөний:"
        ws["V" + str(yy)].border = thin_border
        ws["W" + str(yy)] = "Үгийн:"
        ws["W" + str(yy)].border = thin_border
        ws["X" + str(yy)] = "Өөртөө чиглэсэн:"
        ws["X" + str(yy)].border = thin_border
        ws["Y" + str(yy)] = "Дүрслэх:"
        ws["Y" + str(yy)].border = thin_border
        yy+=1

        ws["P" + str(yy)] = f"{ddoc.id}"
        ws["P" + str(yy)].border = thin_border

        index = chars.index("Q")
        for _, percentage in data.items():
            ws[chars[index] + str(yy)] = f"{percentage}%"
            ws[chars[index] + str(yy)].border = thin_border
            index += 1
        yy += 2

        
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream
        




        
    
    



    



    return None
